import openpyxl
from datetime import datetime, date
from decimal import Decimal
from django.utils.translation import gettext_lazy as _
from django.core.exceptions import ValidationError
from django.db.models import Sum
from django.utils import timezone
from .models import OperationLog, Course, Product, CourseInventory
from accounts.models import Member, Coach

def log_operation(user, module, content):
    """
    Log an operation in the system
    
    Args:
        user: The user performing the operation
        module: The module where the operation occurred (e.g., 'members', 'orders')
        content: Description of the operation
    """
    try:
        OperationLog.objects.create(
            user=user,
            module=module,
            operation_content=content
        )
    except Exception as e:
        # Log error but don't stop the operation
        print(f"Error logging operation: {str(e)}")  # In production, use proper logging


class ExcelProcessor:
    """Utility class for processing Excel files for data import"""
    
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        
    def open_workbook(self):
        """Open the Excel workbook"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            return True
        except Exception as e:
            raise ValidationError(f"無法開啟 Excel 檔案: {str(e)}")
    
    def get_data_rows(self, start_row=2):
        """Get data rows from the worksheet (skip header)"""
        if not self.worksheet:
            raise ValidationError("請先開啟 Excel 檔案")
        
        data_rows = []
        for row in range(start_row, self.worksheet.max_row + 1):
            row_data = []
            for col in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=row, column=col)
                row_data.append(cell.value)
            # Only add non-empty rows
            if any(cell is not None for cell in row_data):
                data_rows.append((row, row_data))
        
        return data_rows
    
    def close_workbook(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()


class MembersImportProcessor(ExcelProcessor):
    """Processor for members import"""
    
    REQUIRED_COLUMNS = [
        '姓名', '手機號碼', '負責教練', '註冊日期', '狀態'
    ]
    
    def validate_data(self, data_rows):
        """Validate members import data"""
        errors = []
        warnings = []
        
        for row_num, row_data in data_rows:
            if len(row_data) < 14:
                errors.append(f"第 {row_num} 行：資料欄位不足（應有 14 欄，實際 {len(row_data)} 欄）")
                continue
            
            # Extract data
            name = row_data[0]
            phone = row_data[1]
            address = row_data[2] if len(row_data) > 2 else ''
            gender = row_data[3] if len(row_data) > 3 else ''
            birthday = row_data[4] if len(row_data) > 4 else None
            email = row_data[5] if len(row_data) > 5 else ''
            line_id = row_data[6] if len(row_data) > 6 else ''
            emergency_contact_name = row_data[7] if len(row_data) > 7 else ''
            emergency_contact_relation = row_data[8] if len(row_data) > 8 else ''
            emergency_contact_phone = row_data[9] if len(row_data) > 9 else ''
            coach_name = row_data[10] if len(row_data) > 10 else ''
            register_date = row_data[11] if len(row_data) > 11 else None
            notes = row_data[12] if len(row_data) > 12 else ''
            status = row_data[13] if len(row_data) > 13 else ''
            
            # Validate required fields
            if not name or not str(name).strip():
                errors.append(f"第 {row_num} 行：姓名為必填欄位")
            
            if not phone or not str(phone).strip():
                errors.append(f"第 {row_num} 行：手機號碼為必填欄位")
            else:
                # Check for duplicate phone
                if Member.objects.filter(phone=str(phone).strip()).exists():
                    errors.append(f"第 {row_num} 行：手機號碼 {phone} 已存在")
            
            if not coach_name or not str(coach_name).strip():
                errors.append(f"第 {row_num} 行：負責教練為必填欄位")
            else:
                # Check coach exists
                if not Coach.objects.filter(name=str(coach_name).strip()).exists():
                    errors.append(f"第 {row_num} 行：教練 {coach_name} 不存在")
            
            if not register_date:
                errors.append(f"第 {row_num} 行：註冊日期為必填欄位")
            else:
                try:
                    if isinstance(register_date, str):
                        register_date = datetime.strptime(register_date, '%Y/%m/%d').date()
                    elif isinstance(register_date, datetime):
                        register_date = register_date.date()
                except:
                    errors.append(f"第 {row_num} 行：註冊日期格式錯誤")
            
            _status_map = {'啟用': 'active', '停用': 'inactive'}
            if not status or not str(status).strip():
                errors.append(f"第 {row_num} 行：狀態為必填欄位")
            elif _status_map.get(str(status).strip(), str(status).strip()) not in ['active', 'inactive']:
                errors.append(f"第 {row_num} 行：狀態必須為「啟用」或「停用」")
        
        return errors, warnings
    
    def import_data(self, data_rows, created_by):
        """Import members data"""
        imported_count = 0
        errors = []
        
        for row_num, row_data in data_rows:
            try:
                # Extract data
                name = str(row_data[0]).strip()
                phone = str(row_data[1]).strip()
                address = str(row_data[2]).strip() if len(row_data) > 2 and row_data[2] else ''
                gender = str(row_data[3]).strip() if len(row_data) > 3 and row_data[3] else ''
                birthday = row_data[4] if len(row_data) > 4 and row_data[4] else None
                email = str(row_data[5]).strip() if len(row_data) > 5 and row_data[5] else ''
                line_id = str(row_data[6]).strip() if len(row_data) > 6 and row_data[6] else ''
                emergency_contact_name = str(row_data[7]).strip() if len(row_data) > 7 and row_data[7] else ''
                emergency_contact_relation = str(row_data[8]).strip() if len(row_data) > 8 and row_data[8] else ''
                emergency_contact_phone = str(row_data[9]).strip() if len(row_data) > 9 and row_data[9] else ''
                coach_name = str(row_data[10]).strip()
                register_date = row_data[11]
                notes = str(row_data[12]).strip() if len(row_data) > 12 and row_data[12] else ''
                status_raw = str(row_data[13]).strip()
                _status_map = {'啟用': 'active', '停用': 'inactive'}
                status = _status_map.get(status_raw, status_raw)

                # Get coach
                coach = Coach.objects.get(name=coach_name)
                
                # Map gender from Chinese to model values
                gender_mapping = {
                    '男': 'M',
                    '女': 'F',
                    '其他': 'O'
                }
                gender = gender_mapping.get(gender, gender)
                
                # Parse dates
                if isinstance(register_date, str):
                    register_date = datetime.strptime(register_date, '%Y/%m/%d').date()
                elif isinstance(register_date, datetime):
                    register_date = register_date.date()
                
                if birthday:
                    if isinstance(birthday, str):
                        birthday = datetime.strptime(birthday, '%Y/%m/%d').date()
                    elif isinstance(birthday, datetime):
                        birthday = birthday.date()
                
                # Create member
                member = Member.objects.create(
                    name=name,
                    phone=phone,
                    address=address,
                    gender=gender,
                    birthday=birthday,
                    email=email,
                    line_id=line_id,
                    emergency_contact_name=emergency_contact_name,
                    emergency_contact_relation=emergency_contact_relation,
                    emergency_contact_phone=emergency_contact_phone,
                    coach=coach,
                    registration_date=register_date,
                    notes=notes,
                    status=status
                )
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f"第 {row_num} 行：導入失敗 - {str(e)}")
        
        return imported_count, errors


class OrdersImportProcessor(ExcelProcessor):
    """Processor for orders import"""
    
    def validate_data(self, data_rows):
        """Validate orders import data"""
        errors = []
        warnings = []
        
        for row_num, row_data in data_rows:
            if len(row_data) < 7:
                errors.append(f"第 {row_num} 行：資料欄位不足")
                continue
            
            # Extract data
            member_phone = row_data[0]
            order_date = row_data[1]
            course_name = row_data[2]
            quantity = row_data[3]
            unit_price = row_data[4]
            payment_method = row_data[5]
            order_status = row_data[6]
            notes = row_data[7] if len(row_data) > 7 else ''
            
            # Validate required fields
            if not member_phone or not str(member_phone).strip():
                errors.append(f"第 {row_num} 行：會員手機號碼為必填欄位")
            else:
                # Check member exists
                if not Member.objects.filter(phone=str(member_phone).strip()).exists():
                    errors.append(f"第 {row_num} 行：會員 {member_phone} 不存在")
            
            if not order_date:
                errors.append(f"第 {row_num} 行：訂單日期為必填欄位")
            else:
                try:
                    if isinstance(order_date, str):
                        order_date = datetime.strptime(order_date, '%Y/%m/%d').date()
                    elif isinstance(order_date, datetime):
                        order_date = order_date.date()
                except:
                    errors.append(f"第 {row_num} 行：訂單日期格式錯誤")
            
            if not course_name or not str(course_name).strip():
                errors.append(f"第 {row_num} 行：課程名稱為必填欄位")
            else:
                # Check course exists
                if not Course.objects.filter(name=str(course_name).strip()).exists():
                    errors.append(f"第 {row_num} 行：課程 {course_name} 不存在")
            
            if not quantity or not str(quantity).strip():
                errors.append(f"第 {row_num} 行：購買數量為必填欄位")
            else:
                try:
                    quantity = int(quantity)
                    if quantity <= 0:
                        errors.append(f"第 {row_num} 行：購買數量必須大於0")
                except:
                    errors.append(f"第 {row_num} 行：購買數量格式錯誤")
            
            if not unit_price or not str(unit_price).strip():
                errors.append(f"第 {row_num} 行：單價為必填欄位")
            else:
                try:
                    unit_price = float(unit_price)
                    if unit_price < 0:
                        errors.append(f"第 {row_num} 行：單價不能為負數")
                except:
                    errors.append(f"第 {row_num} 行：單價格式錯誤")
            
            if not payment_method or not str(payment_method).strip():
                errors.append(f"第 {row_num} 行：付款方式為必填欄位")
            else:
                payment_method = str(payment_method).strip()
                valid_payment_methods = ['cash', 'card', 'transfer', '現金', '刷卡', '轉帳']
                if payment_method not in valid_payment_methods:
                    errors.append(f"第 {row_num} 行：付款方式必須為 '現金', '刷卡', 或 '轉帳'")
            
            if not order_status or not str(order_status).strip():
                errors.append(f"第 {row_num} 行：訂單狀態為必填欄位")
            else:
                order_status = str(order_status).strip()
                valid_order_statuses = ['paid', 'unpaid', 'cancelled', '已付款', '未付款', '已取消']
                if order_status not in valid_order_statuses:
                    errors.append(f"第 {row_num} 行：訂單狀態必須為 '已付款', '未付款', 或 '已取消'")
        
        return errors, warnings
    
    def import_data(self, data_rows, created_by):
        """Import orders data"""
        from .models import Order, OrderItem
        
        imported_count = 0
        errors = []
        
        for row_num, row_data in data_rows:
            try:
                # Extract data
                member_phone = str(row_data[0]).strip()
                order_date = row_data[1]
                course_name = str(row_data[2]).strip()
                quantity = int(row_data[3])
                unit_price = Decimal(str(row_data[4]))
                payment_method = str(row_data[5]).strip()
                order_status = str(row_data[6]).strip()
                notes = str(row_data[7]).strip() if len(row_data) > 7 and row_data[7] else ''
                
                # Get member and course
                member = Member.objects.get(phone=member_phone)
                course = Course.objects.get(name=course_name)
                
                # Map Chinese values to English for database storage
                payment_method_mapping = {
                    '現金': 'cash',
                    '刷卡': 'card', 
                    '轉帳': 'transfer'
                }
                payment_method = payment_method_mapping.get(payment_method, payment_method)
                
                order_status_mapping = {
                    '已付款': 'paid',
                    '未付款': 'unpaid',
                    '已取消': 'cancelled'
                }
                order_status = order_status_mapping.get(order_status, order_status)
                
                # Parse date
                if isinstance(order_date, str):
                    order_date = datetime.strptime(order_date, '%Y/%m/%d').date()
                elif isinstance(order_date, datetime):
                    order_date = order_date.date()
                
                # Create order
                order = Order.objects.create(
                    member=member,
                    order_date=order_date,
                    payment_method=payment_method,
                    order_status=order_status,
                    notes=notes,
                    created_by=created_by
                )
                
                # Create order item
                # Note: OrderItem.save() automatically calls order.calculate_total()
                # and order.save(), so no need to call them manually here.
                OrderItem.objects.create(
                    order=order,
                    item_type='course',
                    item_id=course.id,
                    quantity=quantity,
                    unit_price=unit_price,
                    total_price=quantity * unit_price
                )
                
                # Create course inventory for paid orders
                if order_status == 'paid':
                    CourseInventory.objects.create(
                        member=member,
                        course=course,
                        order=order,
                        initial_quantity=quantity,
                        remaining_quantity=quantity
                    )
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f"第 {row_num} 行：導入失敗 - {str(e)}")
        
        return imported_count, errors


class SessionsImportProcessor(ExcelProcessor):
    """Processor for sessions import"""
    
    def validate_data(self, data_rows):
        """Validate sessions import data"""
        errors = []
        warnings = []

        # Track valid (phone, course) pairs for batch inventory check
        session_counts = {}

        for row_num, row_data in data_rows:
            if len(row_data) < 4:
                errors.append(f"第 {row_num} 行：資料欄位不足（應有至少 4 欄）")
                continue

            # Column order: 手機號碼, 課程名稱, 上課教練, 授課日期與時間, 備註
            member_phone = row_data[0]
            course_name = row_data[1]
            coach_name = row_data[2]
            session_date = row_data[3]

            row_valid = True

            # Validate member phone
            if not member_phone or not str(member_phone).strip():
                errors.append(f"第 {row_num} 行：會員手機號碼為必填欄位")
                row_valid = False
            else:
                if not Member.objects.filter(phone=str(member_phone).strip()).exists():
                    errors.append(f"第 {row_num} 行：會員 {member_phone} 不存在")
                    row_valid = False

            # Validate course name
            if not course_name or not str(course_name).strip():
                errors.append(f"第 {row_num} 行：課程名稱為必填欄位")
                row_valid = False
            else:
                if not Course.objects.filter(name=str(course_name).strip(), status='active').exists():
                    errors.append(f"第 {row_num} 行：課程「{course_name}」不存在或已停用")
                    row_valid = False

            # Validate coach name
            if not coach_name or not str(coach_name).strip():
                errors.append(f"第 {row_num} 行：上課教練為必填欄位")
            else:
                if not Coach.objects.filter(name=str(coach_name).strip()).exists():
                    errors.append(f"第 {row_num} 行：教練「{coach_name}」不存在")

            # Validate session date
            if not session_date:
                errors.append(f"第 {row_num} 行：授課日期與時間為必填欄位")
                row_valid = False
            else:
                try:
                    parsed_date = None
                    if isinstance(session_date, str):
                        for fmt in ('%Y/%m/%d %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d', '%Y-%m-%d'):
                            try:
                                parsed_date = datetime.strptime(session_date, fmt)
                                break
                            except ValueError:
                                continue
                        if parsed_date is None:
                            raise ValueError
                    elif isinstance(session_date, datetime):
                        parsed_date = session_date
                    elif isinstance(session_date, date):
                        parsed_date = datetime.combine(session_date, datetime.min.time())
                    else:
                        raise ValueError

                    if parsed_date.date() > date.today():
                        errors.append(f"第 {row_num} 行：授課日期不能晚於今天")
                        row_valid = False
                except (ValueError, TypeError):
                    errors.append(f"第 {row_num} 行：授課日期與時間格式錯誤（建議：YYYY/MM/DD HH:MM:SS）")
                    row_valid = False

            # Accumulate valid rows for batch inventory check
            if row_valid and member_phone and course_name:
                key = (str(member_phone).strip(), str(course_name).strip())
                session_counts[key] = session_counts.get(key, 0) + 1

        # Batch inventory sufficiency check
        for (phone, course_name), count in session_counts.items():
            try:
                member = Member.objects.get(phone=phone)
                course = Course.objects.get(name=course_name)
                total_remaining = CourseInventory.objects.filter(
                    member=member,
                    course=course,
                    remaining_quantity__gt=0
                ).aggregate(total=Sum('remaining_quantity'))['total'] or 0
                if count > total_remaining:
                    errors.append(
                        f"會員 {phone} 課程「{course_name}」庫存不足：本次導入 {count} 堂，剩餘庫存 {total_remaining} 堂"
                    )
            except (Member.DoesNotExist, Course.DoesNotExist):
                pass  # Already caught in per-row validation

        return errors, warnings
    
    def import_data(self, data_rows, created_by):
        """Import sessions data with automatic inventory deduction"""
        from .models import Session

        imported_count = 0
        errors = []

        for row_num, row_data in data_rows:
            try:
                # Column order: 手機號碼, 課程名稱, 上課教練, 授課日期與時間, 備註
                member_phone = str(row_data[0]).strip()
                course_name = str(row_data[1]).strip()
                coach_name = str(row_data[2]).strip() if row_data[2] else ''
                session_date = row_data[3]
                notes = str(row_data[4]).strip() if len(row_data) > 4 and row_data[4] else ''

                member = Member.objects.get(phone=member_phone)
                course = Course.objects.get(name=course_name)

                # Use specified coach; fall back to member's assigned coach
                coach = member.coach
                if coach_name:
                    try:
                        coach = Coach.objects.get(name=coach_name)
                    except Coach.DoesNotExist:
                        pass

                # Parse datetime
                if isinstance(session_date, str):
                    for fmt in ('%Y/%m/%d %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d', '%Y-%m-%d'):
                        try:
                            session_date = datetime.strptime(session_date, fmt)
                            break
                        except ValueError:
                            continue
                elif isinstance(session_date, date) and not isinstance(session_date, datetime):
                    session_date = datetime.combine(session_date, datetime.min.time())

                if not isinstance(session_date, datetime):
                    session_date = timezone.now()

                if timezone.is_naive(session_date):
                    session_date = timezone.make_aware(session_date)

                # Find the earliest course inventory ordered by order date (per spec)
                course_inventory = CourseInventory.objects.filter(
                    member=member,
                    course=course,
                    remaining_quantity__gt=0
                ).select_related('order').order_by('order__order_date').first()

                if not course_inventory:
                    errors.append(f"第 {row_num} 行：會員 {member_phone} 的課程「{course_name}」庫存不足")
                    continue

                Session.objects.create(
                    member=member,
                    coach=coach,
                    course=course,
                    course_inventory=course_inventory,
                    session_date=session_date,
                    notes=notes,
                    created_by=created_by
                )

                course_inventory.use_quantity(1)
                imported_count += 1

            except Exception as e:
                errors.append(f"第 {row_num} 行：導入失敗 - {str(e)}")

        return imported_count, errors


def create_excel_template(import_type):
    """Create Excel template for import"""
    from django.http import HttpResponse
    from openpyxl.styles import NamedStyle
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    if import_type == 'members':
        # Members template
        headers = [
            '姓名', '手機號碼', '地址', '性別', '生日', 'Email', 'Line ID', 
            '緊急聯絡人姓名', '緊急聯絡人關係', '緊急聯絡人電話', '負責教練', '註冊日期', '備註', '狀態'
        ]
        worksheet.append(headers)
        
        # Add sample data
        sample_data = [
            'Demo Member', '0900000000', 'Demo Address', '男', '1990/01/01',
            'demo@example.com', 'demo_line_id', 'Demo Contact', '家人', '0900000001',
            'Demo Coach', '2024/01/01', 'Demo Note', '啟用'
        ]
        worksheet.append(sample_data)
        
        # Set phone number columns as text format
        # Column B (手機號碼) and Column J (緊急聯絡人電話)
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=2).number_format = '@'  # 手機號碼
            worksheet.cell(row=row, column=10).number_format = '@'  # 緊急聯絡人電話
        
        # Set the entire columns as text format for future data entry
        for col in [2, 10]:  # B and J columns
            for row in range(1, 1000):  # Set format for first 1000 rows
                worksheet.cell(row=row, column=col).number_format = '@'
        
    elif import_type == 'orders':
        # Orders template
        headers = [
            '會員手機號碼', '訂單日期', '課程名稱', '購買數量', '單價', 
            '付款方式', '訂單狀態', '備註'
        ]
        worksheet.append(headers)
        
        # Add sample data
        sample_data = [
            '0900000000', '2024/01/01', 'Demo Course', '10', '1000',
            '現金', '未付款', 'Demo Order Note'
        ]
        worksheet.append(sample_data)
        
        # Set phone number column as text format
        # Column A (會員手機號碼)
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=1).number_format = '@'  # 會員手機號碼
        
        # Set the entire column as text format for future data entry
        for row in range(1, 1000):  # Set format for first 1000 rows
            worksheet.cell(row=row, column=1).number_format = '@'
        
    elif import_type == 'sessions':
        # Sessions template
        headers = [
            '會員手機號碼', '課程名稱', '上課教練', '授課日期與時間', '備註'
        ]
        worksheet.append(headers)

        # Add sample data
        sample_data = [
            '0900000000', 'Demo Course', 'Demo Coach', '2024/01/15 14:30:00', 'Demo Session Note'
        ]
        worksheet.append(sample_data)
        
        # Set phone number column as text format
        # Column A (會員手機號碼)
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=1).number_format = '@'  # 會員手機號碼
        
        # Set the entire column as text format for future data entry
        for row in range(1, 1000):  # Set format for first 1000 rows
            worksheet.cell(row=row, column=1).number_format = '@'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{import_type}_import_template.xlsx"'
    
    workbook.save(response)
    return response
